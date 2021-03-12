import openpyxl

def importNames(filename, logger):
    """
    Try to import names from provided Excel file and return it as a list. If
    errors occur, return an empty list.

    -filename: string of Excel file containing the names to be imported.
    """

    names = []

    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        for i in range(1, sheet.max_row+1):
            name = str(sheet.cell(row=i, column=1).value).strip()
            names.append(name)
    except:
        logger.log("An error occured while trying to process a file.", "error")

    return names
