from openpyxl import Workbook, load_workbook

def exportToExcel(formatted_data, f_out, append=False):
    """
    This function exports well-formatted data into an excel file. If the dataset
    is empty, no file will be created.

    - formatted_data: The data to be exported.
    - f_out: The name of the file to be created.
    """

    # Ensure that the dataset is not empty
    if formatted_data:

        if not append:
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "Αριθμός Εγγραφής"
            sheet["B1"] = "Όνομα"
            sheet["C1"] = "Διεύθυνση"
            sheet["D1"] = "Τηλέφωνο"
            row = 2

        else:
            workbook = load_workbook(f_out+".xlsx")
            sheet = workbook.active
            row = sheet.max_row + 1

        for record in formatted_data:
            sheet["A"+str(row)] = record["id"]
            sheet["B"+str(row)] = record["name"]
            street = ""
            if record["street"]: street = str(record["street"])
            streetNumber = ""
            if record["streetNumber"]: streetNuber = str(record["streetNumber"])
            sheet["C"+str(row)] = street + " " + streetNumber
            sheet["D"+str(row)] = record["phoneNumber"]
            row += 1

        workbook.save(filename=f_out+".xlsx")
        return True

    return False
